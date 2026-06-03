import pandas as pd
import json
from datetime import datetime
from pathlib import Path

import matplotlib.pyplot as plt
from io import BytesIO
from openpyxl.styles import Font
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
import numpy as np


class DataExporter:
    @staticmethod
    def export_all(results, base_name='uv_analysis'):
        """Export results to CSV, JSON, and Excel with embedded histograms"""
        Path('outputs/reports').mkdir(parents=True, exist_ok=True)
        Path('outputs/figures').mkdir(parents=True, exist_ok=True)

        df = DataExporter._results_to_dataframe(results)

        #csv
        csv_path = f'outputs/reports/{base_name}.csv'
        df.to_csv(csv_path, index=False, float_format='%.2f')
        print(f"CSV saved: {csv_path}")
        #JSON
        json_path = f'outputs/reports/{base_name}.json'
        json_data = {
            'metadata': {'timestamp': datetime.now().isoformat()},
            'statistics': df.to_dict('records'),
            'summary': {'total_measurements': len(df)}
        }
        with open(json_path, 'w') as f:
            json.dump(json_data, f, indent=2)
        print(f"✓ JSON saved: {json_path}")
        DataExporter._create_excel_with_histograms(results, df, base_name)

        return df

    @staticmethod
    def _create_excel_with_histograms(results, df, base_name):
        excel_path = f'outputs/reports/{base_name}.xlsx'

        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            #sheet 1: Raw data
            df.to_excel(writer, sheet_name='Data', index=False)

            #sheet 2: Pivot table summary
            pivot = df.pivot_table(
                values='mean_intensity',
                index='timepoint_hours',
                columns='roi_name',
                aggfunc='mean'
            ).round(2)
            pivot.to_excel(writer, sheet_name='Summary')

            #sheet 3: Statistics summary
            stats_summary = df.groupby(['timepoint_hours', 'roi_name']).agg({
                'mean_intensity': 'mean',
                'std_dev': 'mean',
                'min_intensity': 'min',
                'max_intensity': 'max'
            }).round(2)
            stats_summary.to_excel(writer, sheet_name='Detailed Stats')

        from openpyxl import load_workbook
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.utils import get_column_letter

        wb = load_workbook(excel_path)
        hist_sheet = wb.create_sheet("Histograms")
        timepoints = sorted(results.keys())
        current_row = 1

        for idx, time in enumerate(timepoints):
            fig, ax = plt.subplots(figsize=(10, 6))
            for roi_name in results[time].keys():
                intensities = results[time][roi_name]['intensities']
                ax.hist(intensities, bins=50, alpha=0.6, label=roi_name)

            ax.set_xlabel('Intensity (0-255)')
            ax.set_ylabel('Pixel Count')
            ax.set_title(f'Timepoint: {time} hours')
            ax.legend()
            ax.grid(True, alpha=0.3)
            img_data = BytesIO()
            plt.savefig(img_data, format='png', dpi=150, bbox_inches='tight')
            plt.close()
            img_data.seek(0)
            title_cell = hist_sheet.cell(row=current_row, column=1, value=f"Timepoint {time} hours")
            title_cell.font = Font(bold=True, size=14)
            img = XLImage(img_data)
            img.width = 800
            img.height = 500
            img.anchor = f'A{current_row + 1}'
            hist_sheet.add_image(img)

            current_row += 35

        hist_sheet.column_dimensions['A'].width = 30
        wb.save(excel_path)

        print(f"✓ Excel with histograms saved: {excel_path}")
    @staticmethod
    def _results_to_dataframe(results):
        rows = []
        for time in sorted(results.keys()):
            for roi_name, roi_data in results[time].items():
                s = roi_data['stats']
                rows.append({
                    'timepoint_hours': time,
                    'roi_name': roi_name,
                    'min_intensity': s['min'],
                    'max_intensity': s['max'],
                    'mean_intensity': s['mean'],
                    'median_intensity': s['median'],
                    'std_dev': s['std'],
                    'range': s['range'],
                    'pixel_count': s['pixel_count'],
                    'kurtosis': s['kurtosis'],
                    'skewness': s.get('skewness', 0),
                    'quality_score': s.get('quality_score', 0),
                    'quality_rating': s.get('quality_rating', 'N/A'),
                })
        return pd.DataFrame(rows)

    @staticmethod
    def print_statistics(results):
        df = DataExporter._results_to_dataframe(results)
        print("\n" + "=" * 80)
        print("ANALYSIS RESULTS")
        print("=" * 80)
        print(df.round(2).to_string(index=False))